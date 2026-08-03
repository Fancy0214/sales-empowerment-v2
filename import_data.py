import openpyxl
from collections import Counter
import json
import httpx
import time

SUPABASE_URL = "https://hgtxozgpvccgsvslokud.supabase.co"
ANON_KEY = "sb_publishable_9Sc9FFYAqKl2eJUdyP0HmA_w8RdAcKH"
HEADERS = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

# 脱敏映射
SALES_MAP = {
    "刘济超": "销售D",
    "李心雨": "销售M",
    "彭凡珊": "销售F",
}

# 级别映射
GRADE_MAP = {
    "A+级（一级代理）": "A+",
    "B级（大型企业）": "B",
    "C级（中型企业）": "C",
    "D级（小型企业）": "D",
    "E级（微型企业）": "E",
    "F级（个人）": "F",
    "A+": "A+", "B": "B", "C": "C", "D": "D", "E": "E", "F": "F",
}

def parse_grade(val):
    if not val:
        return ""
    val = str(val).strip()
    return GRADE_MAP.get(val, val)

def parse_date(val):
    if not val:
        return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
            try:
                import datetime
                return datetime.datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except:
                continue
        return val
    import datetime
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)

def upload_batch(records):
    """批量插入Supabase，每批最多100条"""
    url = f"{SUPABASE_URL}/rest/v1/institution_data"
    total = len(records)
    success = 0
    for i in range(0, total, 100):
        batch = records[i:i+100]
        payload = json.dumps(batch)
        try:
            resp = httpx.post(url, headers=HEADERS, content=payload, timeout=30)
            if resp.status_code in (200, 201):
                success += len(batch)
                print(f"  ✅ 批次 {i+1}-{i+len(batch)}/{total} 导入成功")
            else:
                print(f"  ❌ 批次 {i+1}-{i+len(batch)}/{total} 失败: {resp.status_code} {resp.text[:200]}")
        except Exception as e:
            print(f"  ❌ 批次 {i+1}-{i+len(batch)}/{total} 异常: {e}")
        time.sleep(0.5)
    return success

def main():
    all_records = []
    
    # ========================
    # File1: 增量 (475条)
    # ========================
    f1_path = "/app/data/所有对话/主对话/用户上传/2025年9月1日-2026年7月31日部门有效及有效作废数据 - 机构市场数据列表(部门)-纯表头数据条数 (5)_1785479326941_0_yqrz.xlsx"
    print("=== 解析 File1 (增量) ===")
    wb1 = openpyxl.load_workbook(f1_path, data_only=True)
    ws1 = wb1.active
    headers1 = [cell.value for cell in ws1[1]]
    col1 = {h: i for i, h in enumerate(headers1)}
    rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
    print(f"  总行数: {len(rows1)}")
    
    # 构建File3查找表（补充出产状态）
    f3_path = "/app/data/所有对话/主对话/用户上传/部门合作中增量机构_1785480622950_0_lr56.xlsx"
    wb3 = openpyxl.load_workbook(f3_path, data_only=True)
    ws3 = wb3.active
    headers3 = [cell.value for cell in ws3[1]]
    col3 = {h: i for i, h in enumerate(headers3)}
    rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
    
    # File3 lookup by 合同潜在代码
    file3_by_potential = {}
    for r in rows3:
        code = r[col3.get('合同潜在代码', 0)]
        if code:
            file3_by_potential[str(code)] = r
    
    f1_count = 0
    for r in rows1:
        sales_real = str(r[col1['签约跟进人员']]) if r[col1['签约跟进人员']] else ""
        sales_person = SALES_MAP.get(sales_real, "其他销售" if sales_real else "")
        
        grade_col = '入学量等级预估' if '入学量等级预估' in col1 else '入学量梯度预估'
        grade_raw = str(r[col1[grade_col]]) if r[col1[grade_col]] else ""
        grade_level = parse_grade(grade_raw)
        
        # 合作状态
        contract_status = str(r[col1['合作状态']]) if r[col1['合作状态']] else ""
        
        # 出产情况
        prod_raw = str(r[col1['出产情况']]) if r[col1['出产情况']] else ""
        production_status = ""
        if "已出产" in prod_raw:
            production_status = "已出产"
        elif "未出产" in prod_raw:
            production_status = "未出产"
        elif contract_status == "合作中" and not production_status:
            production_status = "未出产"
        elif contract_status == "尚未合作" and not production_status:
            production_status = "未出产"
        
        # 尝试从File3补充出产状态
        if not production_status and contract_status == "合作中":
            potential_code = str(r[col1['潜在合同代码']]) if r[col1['潜在合同代码']] else ""
            if potential_code and potential_code in file3_by_potential:
                r3 = file3_by_potential[potential_code]
                prod3 = str(r3[col3['出产情况']]) if r3[col3['出产情况']] else ""
                if "已出产" in prod3:
                    production_status = "已出产"
                elif prod3:
                    production_status = "未出产"
        
        # 日期
        contract_date = parse_date(r[col1.get('正式签约时间', '')] if col1.get('正式签约时间') else None)
        if not contract_date:
            contract_date = parse_date(r[col1.get('合同生效时间', '')] if col1.get('合同生效时间') else None)
        
        # 渠道供给人 (数据提供人员，排除销售)
        provider_real = str(r[col1['数据提供人员']]) if r[col1['数据提供人员']] else ""
        channel_provider = ""
        channel_provider_real = provider_real if provider_real and provider_real not in ["刘济超", "李心雨", "彭凡珊"] else ""
        
        # 市场来源
        market_source = str(r[col1['市场来源一级']]) if r[col1['市场来源一级']] else ""
        
        # 城市/省份
        city = str(r[col1.get('所在市', '')]) if col1.get('所在市') else ""
        province = str(r[col1.get('所在省', '')]) if col1.get('所在省') else ""
        
        # 机构名称
        inst_name = str(r[col1.get('机构名称', '')]) if col1.get('机构名称') else ""
        
        record = {
            "institution_name": inst_name or None,
            "grade_level": grade_level or None,
            "sales_person": sales_person or None,
            "sales_person_real": sales_real or None,
            "market_source_primary": market_source or None,
            "contract_status": contract_status or None,
            "contract_date": contract_date,
            "production_status": production_status or None,
            "production_date": None,
            "data_category": "增量",
            "data_year": "当年度",
            "city": city or None,
            "province": province or None,
            "channel_provider": channel_provider or None,
            "channel_provider_real": channel_provider_real or None,
            "notes": None,
        }
        all_records.append(record)
        f1_count += 1
    
    print(f"  File1 解析完成: {f1_count} 条")
    
    # ========================
    # File2: 存量 (246条)
    # ========================
    f2_path = "/app/data/所有对话/主对话/用户上传/2025年8月31日之前签约合作的增量机构_1785479326942_1_y3tu.xlsx"
    print("\n=== 解析 File2 (存量) ===")
    wb2 = openpyxl.load_workbook(f2_path, data_only=True)
    ws2 = wb2.active
    headers2 = [cell.value for cell in ws2[1]]
    col2 = {h: i for i, h in enumerate(headers2)}
    rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
    print(f"  总行数: {len(rows2)}")
    
    f2_count = 0
    grade_key = '入学量等级预估(等级)' if '入学量等级预估(等级)' in col2 else '入学量等级预估'
    
    for r in rows2:
        sales_real = str(r[col2['签约跟进人员']]) if r[col2['签约跟进人员']] else ""
        sales_person = SALES_MAP.get(sales_real, "其他销售" if sales_real else "")
        
        grade_raw = str(r[col2[grade_key]]) if col2.get(grade_key) is not None and r[col2[grade_key]] else ""
        grade_level = parse_grade(grade_raw)
        
        contract_status = "合作中"  # 存量都是合作中
        
        prod_raw = str(r[col2['出产情况']]) if r[col2['出产情况']] else ""
        production_status = "已出产" if "已出产" in prod_raw else "未出产"
        
        contract_date = parse_date(r[col2.get('合同生效时间', '')] if col2.get('合同生效时间') else None)
        
        provider_real = str(r[col2['数据提供人员']]) if r[col2['数据提供人员']] else ""
        channel_provider_real = provider_real if provider_real and provider_real not in ["刘济超", "李心雨", "彭凡珊"] else ""
        
        market_source = str(r[col2['市场来源一级']]) if r[col2['市场来源一级']] else ""
        city = str(r[col2.get('所在市', '')]) if col2.get('所在市') else ""
        province = str(r[col2.get('所在省', '')]) if col2.get('所在省') else ""
        inst_name = str(r[col2.get('机构名称', '')]) if col2.get('机构名称') else ""
        
        record = {
            "institution_name": inst_name or None,
            "grade_level": grade_level or None,
            "sales_person": sales_person or None,
            "sales_person_real": sales_real or None,
            "market_source_primary": market_source or None,
            "contract_status": contract_status,
            "contract_date": contract_date,
            "production_status": production_status,
            "production_date": None,
            "data_category": "存量",
            "data_year": "8.31前存量",
            "city": city or None,
            "province": province or None,
            "channel_provider": None,
            "channel_provider_real": channel_provider_real or None,
            "notes": None,
        }
        all_records.append(record)
        f2_count += 1
    
    print(f"  File2 解析完成: {f2_count} 条")
    
    # ========================
    # 统计
    # ========================
    print(f"\n=== 总计 ===")
    print(f"总记录数: {len(all_records)} (File1: {f1_count}, File2: {f2_count})")
    
    # 统计签约率
    total = len(all_records)
    signed = sum(1 for r in all_records if r['contract_status'] == '合作中')
    produced = sum(1 for r in all_records if r['production_status'] == '已出产')
    print(f"合作中: {signed} ({signed/total*100:.1f}%)")
    print(f"已出产: {produced} ({produced/total*100:.1f}%)")
    
    # 先清空现有数据
    print("\n=== 清空现有数据 ===")
    url = f"{SUPABASE_URL}/rest/v1/institution_data"
    resp = httpx.delete(url, headers={**HEADERS, "Prefer": "return=minimal"})
    print(f"  清空结果: {resp.status_code}")
    time.sleep(1)
    
    # 批量上传
    print(f"\n=== 开始导入 {len(all_records)} 条记录 ===")
    success = upload_batch(all_records)
    print(f"\n导入完成: {success}/{len(all_records)} 条成功")
    
    # 验证
    print("\n=== 验证 ===")
    resp = httpx.get(f"{url}?select=count", headers={**HEADERS, "Prefer": "count"}, timeout=10)
    count = resp.headers.get('content-range', '').split('/')[-1] if resp.headers.get('content-range') else "unknown"
    print(f"  表中记录数: {count}")

if __name__ == "__main__":
    main()
